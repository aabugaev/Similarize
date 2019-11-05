#!/usr/bin/env python
# coding: utf-8

# In[1]:


import difflib
import pandas as pd


# In[2]:


Similarize_File_Name = str(input("Please give the name of the file with lines to be similirized.\n"))
Similarize_df = pd.read_excel(Similarize_File_Name)


# In[3]:


Similirize_Col_Name = str(input("Please give the name of the column with lines to be similirized.\n"))
#Similirize_Col = Similarize_df[Similirize_Col_Name]


# In[4]:


Find_File_Name = str(input("Please give the name of the file with keys to be found.\n"))
Find_df = pd.read_excel(Find_File_Name)


# In[5]:


Find_Col_Name = str(input("Please give the name of the column.\n"))
Find_Col = Find_df[Find_Col_Name]


# In[6]:


# Создаем временный датафрейм с уникальными строками, которые будем симиларизировать. (нужно, чтобы не дублировать работу)
Similarize_Col_unique_df = pd.DataFrame(Similarize_df[Similirize_Col_Name].unique(),columns = [Similirize_Col_Name])
Similarize_Col_unique_df["Similar"] = ""
Similarize_Col_unique_df["Score"] = ""


# In[7]:


for index, row in Similarize_Col_unique_df.iterrows():
    try:
        #Находим самый подходящий вариант и записываем его в колонку
        match = difflib.get_close_matches(Similarize_Col_unique_df.loc[index,Similirize_Col_Name], Find_Col, n=1)[0]
        Similarize_Col_unique_df.loc[index, "Similar"] = match
        
        #Подсчитвыаем его индекс похожести и записываем в колонку
        Score = difflib.SequenceMatcher(None, Similarize_Col_unique_df.loc[index,Similirize_Col_Name],match).ratio()*100
        Similarize_Col_unique_df.loc[index, "Score"] = Score
        
        print(difflib.get_close_matches(Similarize_Col_unique_df.loc[index,Similirize_Col_Name], Find_Col, n =1))
        print(Score)
    except:
        print("Error")
        pass


# In[8]:


# Записываем датафрейм с уникальными строками - по нему легче сверять
from datetime import datetime
thetime= datetime.now().strftime("%Y%m%d-%H%M%S")
Similarize_Col_unique_df.to_excel("Similar_"+thetime+".xlsx")


# In[9]:


# джоиним датафрейм с уникальными ключами с исходным датафреймом (где они могут повторяться)
#Merged_df = Similarize_df.merge(Similarize_Col_unique_df, how = "left", on = Similirize_Col_Name)


# In[10]:


#Merged_df.to_excel("Similarized_"+Similarize_File_Name+"_to_"+Find_File_Name+".xlsx")


# In[11]:


"""

Без кавычек:

#%load file.py
%%writefile file.py  - в начале блока
%pycat  -
%run file.py
%lsmagic

from IPython.core.interactiveshell import InteractiveShell
InteractiveShell.ast_node_interactivity = "all"

===savetime===
from datetime import datetime
thetime= datetime.now().strftime("%Y%m%d-%H%M%S")

===openpyxl===
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook()
wb_ws = wb.get_active_sheet()

wrwb = Workbook()
wrwb_ws = wrwb.get_active_sheet()

wb.save()

===numpy/pandas===
import pandas as pd
import numpy as np

excel_df = pd.read_excel()
csv_df = pd.read_csv()


df.to_excel()
df.to_csv()

writer = pd.ExcelWriter('',engine='xlsxwriter',options={})
df.to_excel(writer)
writer.save()


====requests/BeautifulSoup===
import requests
from bs4 import BeautifulSoup

page = requests.get("http://yandex.ru")
page.encoding = "windows-1251"
soup = BeautifulSoup(''.join(page.text), "html.parser\"),
soup.findAll("div")


===Files and directories===
import os
FileList = os.listdir()

#if not os.path.exists("Folder"):
#   os.mkdir("Folder") 

"""


# In[ ]:




